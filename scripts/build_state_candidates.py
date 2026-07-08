#!/usr/bin/env python3
"""
build_state_candidates.py — reproducible builder for the STATE-tier candidate data.

Pulls two official Utah feeds and merges them into 2026_candidates_merged.csv:

  1. le.utah.gov/data/legislators.json  — all 104 sitting legislators (75 House +
     29 Senate) with photo/email/phone/social. The incumbent BASE layer.
  2. vote.utah.gov  Candidate-Filing-2026.xlsx — every 2026 filer with a Status
     (Election Candidate / Primary / Out in Convention / …). The 2026 OVERLAY.

State School Board (15) + US House (4) incumbents are small fixed rosters (not in
the legislator feed) and are listed as constants below.

WHY THIS IS RERUNNABLE: before primary certification (July 20 2026) the filing
sheet marks contested nominees as "Primary" — winner unknown. The RESOLVED dict
lets us pin winners we've already confirmed. AFTER certification the sheet flips
winners to "Election Candidate" and drops losers, so a fresh run needs no manual
input — just rerun this and the remaining PRIMARY-PENDING races resolve themselves.

Usage:  python3 scripts/build_state_candidates.py
Deps:   openpyxl  (pip install openpyxl)
"""
import csv, re, ssl, json, urllib.request, io, sys
try:
    import openpyxl
except ImportError:
    sys.exit("Need openpyxl:  pip install openpyxl")

LEG_URL = "https://le.utah.gov/data/legislators.json"
XLSX_URL = "https://vote.utah.gov/wp-content/uploads/2026/01/Candidate-Filing-2026.xlsx"
OUT = "2026_candidates_merged.csv"

# Confirmed primary winners (pre-certification overrides). Key = (DistrictType,
# Office) -> winning surname. Harmless to leave stale: once the xlsx flips these
# to "Election Candidate" post-cert, the overlay already reflects the winner.
RESOLVED = {
    ("congress", "U.S. House District 1"): "mcadams",
    ("congress", "U.S. House District 2"): "moore",
    ("congress", "U.S. House District 3"): "maloy",
    ("house",    "State House District 16"): "stevenson",
    ("senate",   "State Senate District 18"): "fiefia",
}

# State School Board incumbents (schools.utah.gov/board/utah/members). District -> name.
SCHOOL_BOARD = {1:"Jennie Earl",2:"Joseph Kerry",3:"Rod Hall",4:"LeAnn Wood",5:"Sarah Reale",
    6:"Carol Barlow Lear",7:"Erin Longacre",8:"Christina Boggess",9:"Amanda Bollinger",
    10:"Matt Hymas",11:"Cindy Bishop Davis",12:"Cole Kelley",13:"Randy Boothe",14:"Emily Green",15:"Joann Brinton"}
# US House incumbents (name, party, official congressional site). District
# numbers here are only a fallback for incumbents who DON'T appear in the
# filing; anyone running gets their district from the filing (Moore→2,
# Maloy→3, Kennedy→4 under the 2026 court map). Owens is retiring and has no
# valid new-map seat, so he's intentionally omitted — new District 1 simply
# shows candidates with no returning incumbent.
US_HOUSE = {1:("Blake Moore","R","https://blakemoore.house.gov/"),
            2:("Celeste Maloy","R","https://maloy.house.gov/"),
            3:("Mike Kennedy","R","https://mikekennedy.house.gov/")}

COLS = ["Name","Office","DistrictType","District","Party","Incumbent","Status","PhotoURL","Bio",
        "Website","OfficialURL","Email","Phone","Facebook","Instagram","X","VolunteerURL","DonateURL","TopIssues","Active","Order"]
CTX = ssl.create_default_context(); CTX.check_hostname=False; CTX.verify_mode=ssl.CERT_NONE

def fetch(url):
    req=urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0 (precinct-tool build script)"})
    return urllib.request.urlopen(req, timeout=60, context=CTX).read()
def clean(s): return re.sub(r"\s+"," ",str(s or "")).strip()
def toks(n): return [t for t in re.sub(r"[^\w\s]"," ",n.lower()).split() if len(t)>1] or n.lower().split()
def surn(n): return toks(n)[-1]
def fnm(n):  return toks(n)[0]
def clean_office(o): return re.sub(r"\s*\(multi-county\)","",o,flags=re.I).strip()
def blank(): return {c:"" for c in COLS}

NICK={"mike":"michael","jim":"james","bill":"william","will":"william","bob":"robert","rob":"robert",
"dave":"david","tom":"thomas","dan":"daniel","chris":"christopher","matt":"matthew","joe":"joseph",
"nate":"nathan","ben":"benjamin","ken":"kenneth","rick":"richard","steve":"stephen","tony":"anthony",
"jen":"jennifer","liz":"elizabeth","cathy":"catherine"}
def same_person(a,b):
    if a==b or a.startswith(b) or b.startswith(a): return True
    return NICK.get(a)==b or NICK.get(b)==a or (NICK.get(a) and NICK.get(a)==NICK.get(b))

def office_map(o):
    low=o.lower()
    for pat,dt in [(r"u\.?s\.? house district\s*(\d+)","congress"),(r"state senate district\s*(\d+)","senate"),
                   (r"state house district\s*(\d+)","house"),(r"state school board dist?r?ict\s*(\d+)","school_board")]:
        m=re.match(pat,low)
        if m: return dt,m.group(1)
    return None,None

def titlecase(n):
    s=re.sub(r"\b([A-Za-z])([A-Za-z']*)", lambda m:m.group(1).upper()+m.group(2).lower(), n)
    s=re.sub(r"\bMc([a-z])", lambda m:"Mc"+m.group(1).upper(), s)   # Mcadams -> McAdams
    s=re.sub(r"\bO'([a-z])", lambda m:"O'"+m.group(1).upper(), s)   # O'dell  -> O'Dell
    return s

PARTY_FULL={"R":"Republican","D":"Democratic","U":"Unaffiliated","L":"Libertarian","I":"Independent American"}
def full_party(p): return PARTY_FULL.get(str(p).strip(), str(p).strip())

# The legislator feed stores facebook/instagram as bare "@handle"s, but the
# frontend's normalizeUrl points any "@handle" at x.com — expand them to full
# platform URLs here. X/Twitter may stay an @handle (the app handles that).
def social_url(val, domain):
    v=clean(val)
    if not v: return ""
    if v.startswith("@"): return f"https://www.{domain}/{v[1:].split()[0]}"
    # Some feed entries are display names ("Hoang Nguyen"), not links — drop
    # anything that can't be a URL rather than render a broken href.
    if " " in v or "." not in v: return ""
    return v

def build():
    # ---- incumbents (base layer) ----
    leg=json.loads(fetch(LEG_URL).decode("cp1252"))
    leg=leg["legislators"] if isinstance(leg,dict) and "legislators" in leg else leg
    inc=[]
    for r in leg:
        dt="house" if r.get("house")=="H" else "senate"
        # Senate records carry bio as a literal boolean (True) instead of text —
        # only keep bios that are actual prose.
        bio=r.get("bio")
        bio=clean(bio)[:500] if isinstance(bio,str) and len(clean(bio))>20 else ""
        lid=clean(r.get("id"))
        official=(f"https://house.utleg.gov/rep/{lid}/" if dt=="house" else f"https://senate.utah.gov/sen/{lid}/") if lid else ""
        inc.append({"Name":clean(r.get("formatName")),"Office":f"State {'House' if dt=='house' else 'Senate'} District {r.get('district')}",
            "DistrictType":dt,"District":str(r.get("district")).strip(),"Party":full_party(clean(r.get("party"))),"Incumbent":"TRUE",
            "PhotoURL":clean(r.get("image")),"Bio":bio,"OfficialURL":official,"Email":clean(r.get("email")),
            "Phone":clean(r.get("cell") or r.get("workPhone") or r.get("homePhone")),
            "Facebook":social_url(r.get("facebook"),"facebook.com"),
            "Instagram":social_url(r.get("instagram"),"instagram.com"),"X":clean(r.get("twitter"))})
    for d,name in SCHOOL_BOARD.items():
        inc.append({"Name":name,"Office":f"State School Board District {d}","DistrictType":"school_board","District":str(d),"Incumbent":"TRUE"})
    for d,(name,party,official) in US_HOUSE.items():
        inc.append({"Name":name,"Office":f"U.S. House District {d}","DistrictType":"congress","District":str(d),"Party":full_party(party),"Incumbent":"TRUE","OfficialURL":official})

    # ---- overlay (2026 filers, keep general + pending) ----
    wb=openpyxl.load_workbook(io.BytesIO(fetch(XLSX_URL)), read_only=True, data_only=True)
    ws=wb[wb.sheetnames[0]]
    cand=[]
    for row in ws.iter_rows(values_only=True):
        vals=[(str(c).strip() if c is not None else "") for c in row]
        if len(vals)<4 or not vals[0] or vals[0]=="Candidate" or not vals[1]: continue
        dt,dist=office_map(vals[1])
        if not dt: continue
        status=vals[3]
        if status=="Election Candidate": st="GENERAL"
        elif status=="Primary": st="PRIMARY-PENDING"
        else: continue  # dropped: convention/withdrew/disqualified/etc.
        cand.append({"Name":titlecase(vals[0]),"Office":vals[1],"DistrictType":dt,"District":dist,"Party":full_party(vals[2]),"Status":st})

    # ---- merge ----
    from collections import defaultdict
    idx=defaultdict(list)
    for r in inc:
        k=("congress",surn(r["Name"])) if r["DistrictType"]=="congress" else (r["DistrictType"],r["District"],surn(r["Name"]))
        idx[k].append(r)
    used=set(); out=[]
    for c in cand:
        dt,dist=c["DistrictType"],c["District"]
        # apply confirmed primary winners
        rk=(dt,clean_office(c["Office"]))
        if c["Status"]=="PRIMARY-PENDING" and rk in RESOLVED:
            if surn(c["Name"])!=RESOLVED[rk]: continue   # lost
            c={**c,"Status":"GENERAL"}
        k=("congress",surn(c["Name"])) if dt=="congress" else (dt,dist,surn(c["Name"]))
        match=next((r for r in idx.get(k,[]) if same_person(fnm(r["Name"]),fnm(c["Name"]))),None)
        row=blank(); row.update({"Name":c["Name"],"Office":clean_office(c["Office"]),"DistrictType":dt,"District":dist,
            "Party":c["Party"],"Status":c["Status"],"Incumbent":"FALSE","Active":"TRUE","Order":"1"})
        if match:
            used.add(id(match)); row["Incumbent"]="TRUE"; row["Name"]=match["Name"]
            for f in ["PhotoURL","Bio","Website","OfficialURL","Email","Phone","Facebook","Instagram","X"]: row[f]=match.get(f,"")
            if dt=="congress": row["Status"]=c["Status"]+" | district per 2026 filing"
        out.append(row)
    for r in inc:
        if id(r) in used: continue
        row=blank(); row.update(r); row["Incumbent"]="TRUE"; row["Office"]=clean_office(r["Office"]); row["Active"]="TRUE"; row["Order"]="1"
        row["Status"]="INCUMBENT — verify 2026 district (new map)" if r["DistrictType"]=="congress" else "INCUMBENT — not on 2026 ballot"
        out.append(row)

    with open(OUT,"w",newline="") as f:
        w=csv.DictWriter(f,fieldnames=COLS); w.writeheader(); w.writerows(out)
    from collections import Counter
    st=Counter(o["Status"].split(" |")[0].split(" —")[0] for o in out)
    print(f"wrote {OUT}: {len(out)} rows | incumbents {sum(1 for o in out if o['Incumbent']=='TRUE')} "
          f"| {dict(st)}")

if __name__=="__main__":
    build()
